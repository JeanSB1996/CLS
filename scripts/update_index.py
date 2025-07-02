#!/usr/bin/env python3
"""Update Excel index with document descriptions.

This script scans a document directory, extracts simple information from PDF
files and updates ``Libro7.xlsx`` with the file name, relative path and a
brief description.
"""

from __future__ import annotations
import argparse
import os
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyPDF2 import PdfReader


def extract_pdf_description(path: Path) -> str:
    """Return a short description for a PDF file."""
    try:
        reader = PdfReader(str(path))
        if reader.pages:
            text = reader.pages[0].extract_text() or ""
        else:
            text = ""
    except Exception:
        return "(no legible)"

    text_low = text.lower()
    # Invoice detection
    if "factura" in text_low or "invoice" in text_low:
        match = re.search(r"proveedor[:\s]+([\w\s]+)", text_low)
        provider = match.group(1).strip() if match else ""
        return f"Factura {provider}".strip()
    # Certificate detection
    if "certificado" in text_low:
        match = re.search(r"emitid[oa] por[:\s]+([\w\s]+)", text_low)
        issuer = match.group(1).strip() if match else ""
        return f"Certificado {issuer}".strip()
    # Fallback first line
    first_line = text.strip().splitlines()[0] if text else ""
    return first_line[:80]


def ensure_headers(ws) -> None:
    headers = ["Name", "Path", "Description"]
    if ws.max_row == 0:
        ws.append(headers)
    else:
        existing = [c.value for c in ws[1]]
        for idx, title in enumerate(headers, 1):
            if idx > len(existing) or existing[idx-1] != title:
                ws.cell(row=1, column=idx, value=title)
    # adjust column width for readability
    for i, w in enumerate((40, 60, 60), 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def update_index(doc_dir: Path, excel_path: Path) -> None:
    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
    ws = wb.active
    ensure_headers(ws)

    existing = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        path_cell = row[1]
        if path_cell.value:
            existing[path_cell.value] = row

    for file in doc_dir.rglob('*'):
        if not file.is_file():
            continue
        rel_path = str(file.relative_to(doc_dir))
        if rel_path in existing:
            row = existing[rel_path]
        else:
            row = ws.append([file.name, rel_path, ""])
            row = ws[ws.max_row]

        description = row[2].value
        if description:
            continue  # keep existing description

        if file.suffix.lower() == '.pdf':
            desc = extract_pdf_description(file)
        else:
            desc = f"Documento {file.suffix.lower()}"
        row[2].value = desc

    wb.save(excel_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Update document index")
    parser.add_argument('--docs', default='Cleantech - Atacama Salt Lakes SpA',
                        help='Directory with documents')
    parser.add_argument('--excel', default='Libro7.xlsx',
                        help='Excel workbook to update')
    args = parser.parse_args()

    doc_dir = Path(args.docs)
    excel_path = Path(args.excel)
    if not doc_dir.is_dir():
        raise SystemExit(f"Document directory {doc_dir} not found")
    update_index(doc_dir, excel_path)


if __name__ == '__main__':
    main()
